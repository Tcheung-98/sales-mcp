"""Fixtures and tests for Logic Guide engine (PI-2760 / I2)."""

import io
from datetime import date

import pytest
from openpyxl import Workbook

from ingestion.gtm_ideation_catalog import GtmIdeationCatalog
from ingestion.inventory_workbook import InventoryWorkbook
from ingestion.logic_guide.engine import LogicGuideEngine
from ingestion.schema import DiscoverySchema


def _gtm_xlsx() -> bytes:
    wb = Workbook()
    ws_cat = wb.active
    ws_cat.title = "Product Category"
    ws_cat.append(["Title", "Description"])
    ws_cat.append(["Newsletter", "Inbox"])
    ws_prod = wb.create_sheet("Product Tags")
    ws_prod.append(
        ["Product Category", "Product Name", "GTM TAGS", "Deck Path", "Slide #"]
    )
    rows = [
        (
            "Newsletters",
            "CEO Daily",
            "ceo, c-suite, leadership",
            "Fortune_Newsletters_2026.pptx",
            "3",
        ),
        (
            "Newsletters",
            "Term Sheet",
            "dealmakers, venture capital, M&A",
            "Fortune_Newsletters_2026.pptx",
            "4",
        ),
        (
            "Digital Ads/Programmatic",
            "Crown Unit",
            "high-impact, digital display",
            "Fortune_High_Impact_2026.pptx",
            "5",
        ),
        (
            "Digital Ads/Programmatic",
            "Scroller Unit",
            "high-impact, scroller",
            "Fortune_High_Impact_2026.pptx",
            "6",
        ),
        (
            "Branded Content",
            "Long-Form Article",
            "written, article, thought leadership",
            "FBS_Content_2026.pptx",
            "9",
        ),
        (
            "Print",
            "Full Page",
            "print, magazine",
            "FortuneAI_DeckTemplate",
            "20",
        ),
    ]
    for row in rows:
        ws_prod.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inventory_xlsx() -> bytes:
    wb = Workbook()
    ws_prod = wb.active
    ws_prod.title = "Products"
    ws_prod.append(
        [
            "Product / Placement",
            "Product Type",
            "Cadence",
            "Launch",
            "Mon",
            "Tue",
            "Wed",
            "Thu",
            "Fri",
            "Sat",
            "Sun",
        ]
    )
    ws_prod.append(
        [
            "CEO Daily",
            "Newsletter",
            "M-F",
            date(2026, 1, 1),
            "✓",
            "✓",
            "✓",
            "✓",
            "✓",
            "",
            "",
        ]
    )
    ws_inv = wb.create_sheet("Inventory")
    ws_inv.append(
        [
            "Date",
            "Day",
            "Week",
            "Month",
            "Product Type",
            "Product / Placement",
            "Status",
        ]
    )
    ws_inv.append(
        [date(2026, 9, 1), "Tue", 36, "Sep", "Newsletter", "CEO Daily", "Available"]
    )
    ws_price = wb.create_sheet("Pricing + Benchmarks")
    ws_price.append(["Section", "Product", "Pricing", "Benchmarks"])
    pricing = [
        ("NEWSLETTERS", "CEO Daily", "$5,000/day"),
        ("NEWSLETTERS", "Term Sheet", "$6,000/day"),
        ("DIGITAL MEDIA", "Crown Unit", "$25,000/day"),
        ("DIGITAL MEDIA", "Scroller Unit", "$15,000/day"),
        ("BRANDED CONTENT", "Long-Form Article", "$60,000"),
        ("PRINT", "Full Page", "$35,000"),
    ]
    for section, product, rate in pricing:
        ws_price.append([section, product, rate, ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def engine() -> LogicGuideEngine:
    gtm = GtmIdeationCatalog.from_xlsx_bytes(_gtm_xlsx())
    inventory = InventoryWorkbook.from_xlsx_bytes(_inventory_xlsx())
    return LogicGuideEngine(gtm, inventory)


def _discovery(**overrides) -> DiscoverySchema:
    base = {
        "company_name": "Acme Corp",
        "industry": "Technology",
        "budgets": [{"amount": 150_000.0, "label": "Primary"}],
        "flight_dates": {"start": "2026-09-01", "end": "2026-12-31"},
        "campaign_goal": "Drive awareness",
        "targeting_details": "CEO and c-suite leadership audience",
        "kpis": ["Awareness"],
        "kpi_details": "Lift awareness",
        "campaign_narrative": "Leadership narrative",
        "preferred_platforms_products": [
            "Newsletters",
            "Digital Ads/Programmatic",
            "Branded Content",
        ],
        "additional_rfp_details": "",
        "client_logo": "https://example.com/logo.png",
    }
    base.update(overrides)
    return DiscoverySchema.model_validate(base)


def test_conference_platform_escalates(engine: LogicGuideEngine):
    discovery = _discovery(
        preferred_platforms_products=["Conference Sponsorship/Media", "Newsletters"]
    )
    result = engine.propose(discovery)
    assert result.requires_gtm_escalation
    assert not result.tiers


def test_proposes_funded_tier(engine: LogicGuideEngine):
    result = engine.propose(_discovery())
    assert not result.requires_gtm_escalation
    assert len(result.tiers) == 1
    tier = result.tiers[0]
    names = {p.name for p in tier.products}
    assert "CEO Daily" in names
    assert "Crown Unit" in names
    assert tier.total <= tier.budget_target or result.notes


def test_multi_tier_subset(engine: LogicGuideEngine):
    discovery = _discovery(
        budgets=[
            {"amount": 200_000.0, "label": "High"},
            {"amount": 80_000.0, "label": "Low"},
        ]
    )
    result = engine.propose(discovery)
    assert len(result.tiers) == 2
    high, low = result.tiers
    assert high.total >= low.total
    assert {p.name for p in low.products}.issubset({p.name for p in high.products})


def test_unavailable_product_dropped(engine: LogicGuideEngine):
    discovery = _discovery(
        flight_dates={"start": "2026-09-01", "end": "2026-09-01"},
    )
    result = engine.propose(discovery)
    # CEO Daily available on 2026-09-01 in fixture
    assert "CEO Daily" not in result.unavailable_products
