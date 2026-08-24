"""Representative Logic Guide workbook fixtures (PI-2760 merge blockers)."""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook

from ingestion.gtm_ideation_catalog import GtmIdeationCatalog
from ingestion.inventory_workbook import InventoryWorkbook
from ingestion.logic_guide.engine import LogicGuideEngine

# --- GTM Product Tags rows: (category, name, tags, deck, slide) ---
REPRESENTATIVE_GTM_ROWS: list[tuple] = [
    (
        "Newsletters",
        "CEO Daily",
        "ceo, c-suite, leadership",
        "Fortune_Newsletters_2026.pptx",
        "3",
    ),
    (
        "Newsletters",
        "Term Sheet",
        "dealmakers, venture capital, M&A, finance",
        "Fortune_Newsletters_2026.pptx",
        "4",
    ),
    (
        "Vodcasts",
        "Term Sheet",
        "dealmakers, venture capital, M&A, finance",
        "Fortune_Premium_Video_2026.pptx",
        "2",
    ),
    (
        "Vodcasts",
        "Fortune Tech",
        "tech, technology, artificial intelligence",
        "Fortune_Premium_Video_2026.pptx",
        "5",
    ),
    (
        "Digital Ads/Programmatic",
        "Crown Unit",
        "high-impact, digital display, any audience",
        "Fortune_High_Impact_2026.pptx",
        "5",
    ),
    (
        "Digital Ads/Programmatic",
        "Scroller Unit",
        "high-impact, scroller, digital display",
        "Fortune_High_Impact_2026.pptx",
        "6",
    ),
    (
        "Digital Ads/Programmatic",
        "LinkedIn BrandLink",
        "linkedin, social, b2b",
        "Fortune_High_Impact_2026.pptx",
        "7",
    ),
    (
        "Digital Ads/Programmatic",
        "Paid Social Video",
        "social, paid social, instagram",
        "Fortune_High_Impact_2026.pptx",
        "8",
    ),
    (
        "Digital Ads/Programmatic",
        "In-Banner Streaming Video",
        "livestream, live stream, streaming",
        "Fortune_High_Impact_2026.pptx",
        "9",
    ),
    (
        "Branded Content",
        "Long-Form Article",
        "written, article, thought leadership",
        "FBS_Content_2026.pptx",
        "9",
    ),
    (
        "Branded Content",
        "Executive Q&A (Remote)",
        "video, executive interview, profile",
        "FBS_Content_2026.pptx",
        "10",
    ),
    (
        "Branded Content",
        "Documentary-Style Video",
        "video, documentary, film, mission",
        "FBS_Content_2026.pptx",
        "11",
    ),
    (
        "Branded Content",
        "Long-Form Q&A Article",
        "written, profile, executive, article",
        "FBS_Content_2026.pptx",
        "12",
    ),
    (
        "Print",
        "Full Page",
        "print, magazine, full page",
        "FortuneAI_DeckTemplate",
        "20",
    ),
]

# (section, product, pricing)
REPRESENTATIVE_PRICING_ROWS: list[tuple[str, str, str]] = [
    ("NEWSLETTERS", "CEO Daily", "$5,000/day"),
    # Single Term Sheet row — pricing sheet is keyed by product name only.
    ("NEWSLETTERS", "Term Sheet", "$6,000/day"),
    ("VODCASTS", "Fortune Tech", "$7,000/day"),
    ("DIGITAL MEDIA", "Crown Unit", "$25,000"),
    ("DIGITAL MEDIA", "Scroller Unit", "$15,000"),
    ("DIGITAL MEDIA", "LinkedIn BrandLink", "$50,000"),
    ("DIGITAL MEDIA", "Paid Social Video", "$25,000"),
    ("DIGITAL MEDIA", "In-Banner Streaming Video", "$20,000"),
    ("BRANDED CONTENT", "Long-Form Article", "$60,000"),
    ("BRANDED CONTENT", "Executive Q&A (Remote)", "$75,000"),
    ("BRANDED CONTENT", "Documentary-Style Video", "$90,000"),
    ("BRANDED CONTENT", "Long-Form Q&A Article", "$65,000"),
    ("PRINT", "Full Page", "$35,000"),
]

# Inventory Products tab rows gated for availability checks
REPRESENTATIVE_INVENTORY_PRODUCTS: list[tuple] = [
    (
        "CEO Daily",
        "Newsletter",
        "M-F",
        date(2026, 1, 1),
        "✓",
        "✓",
        "✓",
        "✓",
        "✓",
        "",
        "",
    ),
    (
        "Term Sheet",
        "Newsletter",
        "M-F",
        date(2026, 1, 1),
        "✓",
        "✓",
        "✓",
        "✓",
        "✓",
        "",
        "",
    ),
]

# (date, day, week, month, type, product, status)
REPRESENTATIVE_INVENTORY_SLOTS: list[tuple] = [
    (date(2026, 9, 1), "Tue", 36, "Sep", "Newsletter", "CEO Daily", "Available"),
    (date(2026, 9, 1), "Tue", 36, "Sep", "Newsletter", "Term Sheet", "Available"),
]


def build_workbook_bytes(
    *,
    gtm_rows: list[tuple] | None = None,
    pricing_rows: list[tuple[str, str, str]] | None = None,
    inventory_products: list[tuple] | None = None,
    inventory_slots: list[tuple] | None = None,
) -> bytes:
    gtm_rows = gtm_rows if gtm_rows is not None else REPRESENTATIVE_GTM_ROWS
    pricing_rows = (
        pricing_rows if pricing_rows is not None else REPRESENTATIVE_PRICING_ROWS
    )
    inventory_products = (
        inventory_products
        if inventory_products is not None
        else REPRESENTATIVE_INVENTORY_PRODUCTS
    )
    inventory_slots = (
        inventory_slots if inventory_slots is not None else REPRESENTATIVE_INVENTORY_SLOTS
    )

    wb = Workbook()
    ws_cat = wb.active
    ws_cat.title = "Product Category"
    ws_cat.append(["Title", "Description"])
    ws_cat.append(["Newsletter", "Inbox"])
    ws_cat.append(["Print", "Print ads"])

    ws_prod = wb.create_sheet("Product Tags")
    ws_prod.append(
        ["Product Category", "Product Name", "GTM TAGS", "Deck Path", "Slide #"]
    )
    for row in gtm_rows:
        ws_prod.append(list(row))

    ws_inv_prod = wb.create_sheet("Products")
    ws_inv_prod.append(
        [
            "Product / Placement",
            "Product Type",
            "Cadence",
            "Launch",
            "Mon",
            "Tue",
            "Wed",
            "Thu",
            "Fri",
            "Sat",
            "Sun",
        ]
    )
    for row in inventory_products:
        ws_inv_prod.append(list(row))

    ws_inv = wb.create_sheet("Inventory")
    ws_inv.append(
        [
            "Date",
            "Day",
            "Week",
            "Month",
            "Product Type",
            "Product / Placement",
            "Status",
        ]
    )
    for row in inventory_slots:
        ws_inv.append(list(row))

    ws_price = wb.create_sheet("Pricing + Benchmarks")
    ws_price.append(
        [
            "Section",
            "Product",
            "Pricing",
            "Benchmarks",
            "Subscribers",
            "Est. Imps",
            "Conference Date",
            "Line Last Updated",
        ]
    )
    for section, product, rate in pricing_rows:
        ws_price.append([section, product, rate, "", "", "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_representative_engine() -> LogicGuideEngine:
    data = build_workbook_bytes()
    return LogicGuideEngine(
        GtmIdeationCatalog.from_xlsx_bytes(data),
        InventoryWorkbook.from_xlsx_bytes(data),
    )


def base_discovery_fields(**overrides) -> dict:
    base = {
        "company_name": "Acme Corp",
        "industry": "Technology",
        "budgets": [{"amount": 250_000.0, "label": "Primary"}],
        "flight_dates": {"start": "2026-09-01", "end": "2026-12-31"},
        "campaign_goal": "Drive awareness among enterprise buyers",
        "targeting_details": "CEO and c-suite leadership",
        "kpis": ["Awareness"],
        "kpi_details": "Lift brand awareness",
        "campaign_narrative": "Enterprise finance modernization",
        "preferred_platforms_products": [
            "Newsletters",
            "Digital Ads/Programmatic",
        ],
        "additional_rfp_details": "",
        "client_logo": "https://example.com/logo.png",
    }
    base.update(overrides)
    return base
