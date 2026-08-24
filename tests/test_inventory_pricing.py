"""Tests for Pricing + Benchmarks loader (Chunk D)."""

import io

import pytest
from openpyxl import Workbook

from ingestion.inventory_pricing import (
    InventoryPricing,
    extract_dollar_amounts,
    load_inventory_pricing_from_s3,
)
from ingestion.inventory_workbook import InventoryWorkbook


def _pricing_xlsx(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing + Benchmarks"
    ws.append(
        [
            "Section",
            "Product",
            "Pricing",
            "Benchmarks",
            "Subscribers",
            "Est. Imps",
            "Conference Date",
            "Line Last Updated",
        ]
    )
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SAMPLE_ROWS = [
    (
        "NEWSLETTERS",
        "Term Sheet",
        "$6,000/day; $24,000/month, 4-day minimum in same quarter",
        "UOR: 42.98% | CTR: 0.02%–0.08%",
        "107,135",
        "",
        "",
        "7/22/2026",
    ),
    (
        "DIGITAL MEDIA",
        "Homepage + First Impression Takeover with Crown",
        "$25,000/day",
        "CTR: 0.05%–0.12%",
        "",
        "300K/day",
        "",
        "",
    ),
    (
        "BRANDED CONTENT (Fortune Brand Studio)",
        "Long-Form Article",
        "$60,000",
        "",
        "",
        "",
        "",
        "7/22/2026",
    ),
    (
        "PRINT",
        "Full Page",
        "",
        "",
        "",
        "",
        "",
        "",
    ),
]


@pytest.fixture
def pricing() -> InventoryPricing:
    return InventoryPricing.from_xlsx_bytes(_pricing_xlsx(_SAMPLE_ROWS))


def test_extract_dollar_amounts():
    assert extract_dollar_amounts("$25,000/day") == (25000.0,)
    assert extract_dollar_amounts(
        "$6,000/day; $24,000/month, 4-day minimum"
    ) == (24000.0, 6000.0)


def test_lookup_verbatim_pricing(pricing: InventoryPricing):
    row = pricing.lookup("Term Sheet")
    assert row.section == "NEWSLETTERS"
    assert "$6,000/day" in row.pricing
    assert row.subscribers == "107,135"


def test_primary_amount(pricing: InventoryPricing):
    assert pricing.primary_amount("Long-Form Article") == 60000.0
    assert pricing.primary_amount("Term Sheet") == 24000.0


def test_lookup_missing_raises(pricing: InventoryPricing):
    with pytest.raises(ValueError, match="No Pricing \\+ Benchmarks row"):
        pricing.lookup("Missing Product")


def test_unpriced_row_excluded_from_lookup():
    data = _pricing_xlsx(
        [
            ("PRINT", "Full Page", "", "", "", "", "", ""),
            ("PRINT", "Cover", "$150,000", "", "", "", "", ""),
        ]
    )
    catalog = InventoryPricing.from_xlsx_bytes(data)
    with pytest.raises(ValueError, match="No Pricing"):
        catalog.lookup("Full Page")
    assert catalog.lookup("Cover").pricing == "$150,000"


def test_missing_sheet_raises():
    wb = Workbook()
    wb.active.title = "Products"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="Pricing \\+ Benchmarks"):
        InventoryPricing.from_xlsx_bytes(buf.getvalue())


def test_workbook_combined_loader():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(
        [
            "Product / Placement",
            "Product Type",
            "Cadence",
            "Launch",
            "Mon",
            "Tue",
            "Wed",
            "Thu",
            "Fri",
            "Sat",
            "Sun",
        ]
    )
    ws.append(
        [
            "CEO Daily",
            "Newsletter",
            "M-F",
            "2026-01-01",
            "✓",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws_inv = wb.create_sheet("Inventory")
    ws_inv.append(
        [
            "Date",
            "Day",
            "Week",
            "Month",
            "Product Type",
            "Product / Placement",
            "Status",
        ]
    )
    ws_inv.append(
        ["2026-09-01", "Tue", 36, "Sep", "Newsletter", "CEO Daily", "Available"]
    )
    ws_price = wb.create_sheet("Pricing + Benchmarks")
    ws_price.append(
        [
            "Section",
            "Product",
            "Pricing",
            "Benchmarks",
            "Subscribers",
            "Est. Imps",
            "Conference Date",
            "Line Last Updated",
        ]
    )
    ws_price.append(
        ["NEWSLETTERS", "CEO Daily", "$5,000/day", "", "100,000", "", "", ""]
    )
    buf = io.BytesIO()
    wb.save(buf)

    book = InventoryWorkbook.from_xlsx_bytes(buf.getvalue())
    assert book.calendar.products.is_inventory_gated("CEO Daily")
    assert book.pricing.lookup("CEO Daily").pricing == "$5,000/day"


def test_live_s3_pricing():
    import os

    import boto3

    bucket = os.environ.get("S3_SNAPSHOT_BUCKET")
    if not bucket:
        pytest.skip("S3_SNAPSHOT_BUCKET not set")
    try:
        pricing = load_inventory_pricing_from_s3(boto3.client("s3"), bucket)
    except Exception as exc:
        pytest.skip(f"Inventory calendar not on S3 yet: {exc}")
    assert len(pricing.product_names) >= 5
    term = pricing.lookup("Term Sheet")
    assert "$" in term.pricing
