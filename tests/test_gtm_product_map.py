"""Tests for exact GTM Product Tags → Deck Path / Slide # mapping."""

import io

import pytest
from openpyxl import Workbook

from ingestion.gtm_product_map import (
    GtmProductMap,
    normalize_category,
    normalize_deck_filename,
    product_deck_s3_key,
)


def _xlsx_bytes(rows: list[tuple]) -> bytes:
    wb = Workbook()
    # Default sheet unused; Product Tags is the SoT tab.
    ws_default = wb.active
    ws_default.title = "Audience Data"
    ws = wb.create_sheet("Product Tags")
    ws.append(
        ["Product Category", "Product Name", "GTM TAGS", "Deck Path", "Slide #"]
    )
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SAMPLE_ROWS = [
    (
        "Newsletters",
        "Term Sheet",
        "finance",
        "Fortune_Newsletters_2026.pptx",
        "4",
    ),
    (
        "Vodcasts",
        "Term Sheet",
        "video",
        "Fortune_Premium_Video_2026.pptx",
        "2",
    ),
    (
        "Newsletters",
        "CEO Daily",
        "ceo",
        "Fortune_Newsletters_2026.pptx",
        "3",
    ),
    (
        "Branded Content",
        "Executive Q&A",
        "full tags",
        "FBS_Content_Offerings_2026_MKTG.pptx",
        "9",
    ),
    (
        "Branded Content",
        "Executive Q&A",
        "bare name",
        "FBS_Content_Offerings_2026_MKTG.pptx",
        "9",
    ),
    (
        "Print",
        "Full Page",
        "print",
        "FortuneAI_DeckTemplate",
        20,
    ),
]


@pytest.fixture
def product_map() -> GtmProductMap:
    return GtmProductMap.from_xlsx_bytes(_xlsx_bytes(_SAMPLE_ROWS))


def test_normalize_category_aliases():
    assert normalize_category("Newsletter") == "Newsletters"
    assert normalize_category("Digital Media") == "Digital Ads/Programmatic"
    assert normalize_category("Branded Content") == "Branded Content"


def test_normalize_deck_filename_adds_pptx():
    assert normalize_deck_filename("FortuneAI_DeckTemplate") == "FortuneAI_DeckTemplate.pptx"
    assert (
        normalize_deck_filename("Fortune_Newsletters_2026.pptx")
        == "Fortune_Newsletters_2026.pptx"
    )


def test_product_deck_s3_key():
    assert (
        product_deck_s3_key("Fortune_Newsletters_2026.pptx")
        == "product-decks/Fortune_Newsletters_2026.pptx"
    )


def test_lookup_unique_name(product_map: GtmProductMap):
    ref = product_map.lookup("CEO Daily", "Newsletter")
    assert ref.deck_path == "Fortune_Newsletters_2026.pptx"
    assert ref.slide_number == 3
    assert ref.product_name == "CEO Daily"


def test_lookup_disambiguates_by_category(product_map: GtmProductMap):
    news = product_map.lookup("Term Sheet", "Newsletter")
    assert news.deck_path == "Fortune_Newsletters_2026.pptx"
    assert news.slide_number == 4

    video = product_map.lookup("Term Sheet", "Vodcasts")
    assert video.deck_path == "Fortune_Premium_Video_2026.pptx"
    assert video.slide_number == 2


def test_lookup_dedupes_identical_duplicate_rows(product_map: GtmProductMap):
    ref = product_map.lookup("Executive Q&A", "Branded Content")
    assert ref.slide_number == 9
    assert ref.deck_path == "FBS_Content_Offerings_2026_MKTG.pptx"


def test_lookup_print_slide_number_as_int(product_map: GtmProductMap):
    ref = product_map.lookup("Full Page", "Print")
    assert ref.deck_path == "FortuneAI_DeckTemplate"
    assert ref.slide_number == 20


def test_lookup_missing_product_fails_loud(product_map: GtmProductMap):
    with pytest.raises(ValueError, match="No GTM Product Tags row for product 'Missing'"):
        product_map.lookup("Missing", "Newsletter")


def test_lookup_ambiguous_without_category_fails(product_map: GtmProductMap):
    with pytest.raises(ValueError, match="Ambiguous GTM Product Tags match"):
        product_map.lookup("Term Sheet")


def test_lookup_wrong_category_fails(product_map: GtmProductMap):
    with pytest.raises(ValueError, match="No GTM Product Tags row for product 'CEO Daily'"):
        product_map.lookup("CEO Daily", "Vodcasts")


def test_missing_sheet_raises():
    wb = Workbook()
    wb.active.title = "Other"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="missing 'Product Tags'"):
        GtmProductMap.from_xlsx_bytes(buf.getvalue())


def test_empty_deck_path_raises():
    data = _xlsx_bytes(
        [("Newsletters", "Broken", "tags", "", "1")]
    )
    with pytest.raises(ValueError, match="empty Deck Path"):
        GtmProductMap.from_xlsx_bytes(data)
