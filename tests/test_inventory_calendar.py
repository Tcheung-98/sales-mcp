"""Tests for inventory calendar loaders (Products + Inventory tabs)."""

import io
from datetime import date

import pytest
from openpyxl import Workbook

from ingestion.inventory_calendar import (
    InventoryAvailability,
    InventoryCalendar,
    InventoryGateResult,
    InventoryProductRegistry,
    load_inventory_calendar_from_s3,
)


def _products_xlsx(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(
        [
            "Product / Placement",
            "Product Type",
            "Cadence",
            "Launch",
            "Mon",
            "Tue",
            "Wed",
            "Thu",
            "Fri",
            "Sat",
            "Sun",
        ]
    )
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inventory_xlsx(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(
        [
            "Date",
            "Day",
            "Week",
            "Month",
            "Product Type",
            "Product / Placement",
            "Status",
            "Sponsor",
        ]
    )
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _combined_xlsx(product_rows: list[tuple], inventory_rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws_prod = wb.active
    ws_prod.title = "Products"
    ws_prod.append(
        [
            "Product / Placement",
            "Product Type",
            "Cadence",
            "Launch",
            "Mon",
            "Tue",
            "Wed",
            "Thu",
            "Fri",
            "Sat",
            "Sun",
        ]
    )
    for row in product_rows:
        ws_prod.append(list(row))
    ws_inv = wb.create_sheet("Inventory")
    ws_inv.append(
        [
            "Date",
            "Day",
            "Week",
            "Month",
            "Product Type",
            "Product / Placement",
            "Status",
            "Sponsor",
        ]
    )
    for row in inventory_rows:
        ws_inv.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_PRODUCT_ROWS = [
    (
        "CEO Daily",
        "Newsletter",
        "M-F, 6:30 a.m.",
        date(2026, 1, 1),
        "✓",
        "✓",
        "✓",
        "✓",
        "✓",
        "",
        "",
    ),
    (
        "Fortune List",
        "Fortune List",
        "Annual (launch→EOY)",
        date(2026, 4, 1),
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ),
]

_INVENTORY_ROWS = [
    (date(2026, 9, 1), "Tue", 36, "Sep", "Newsletter", "CEO Daily", "Available", ""),
    (date(2026, 9, 2), "Wed", 36, "Sep", "Newsletter", "CEO Daily", "Sold", "Acme"),
    (date(2026, 9, 3), "Thu", 36, "Sep", "Newsletter", "CEO Daily", "Available", ""),
    (date(2026, 10, 1), "Thu", 40, "Oct", "Newsletter", "CEO Daily", "Held", ""),
]


@pytest.fixture
def registry() -> InventoryProductRegistry:
    return InventoryProductRegistry.from_xlsx_bytes(_products_xlsx(_PRODUCT_ROWS))


@pytest.fixture
def availability() -> InventoryAvailability:
    return InventoryAvailability.from_xlsx_bytes(_inventory_xlsx(_INVENTORY_ROWS))


@pytest.fixture
def calendar() -> InventoryCalendar:
    return InventoryCalendar.from_xlsx_bytes(
        _combined_xlsx(_PRODUCT_ROWS, _INVENTORY_ROWS)
    )


def test_registry_gated_vs_ungated(registry: InventoryProductRegistry):
    assert registry.is_inventory_gated("CEO Daily")
    assert not registry.is_inventory_gated("Run of Fortune Display")
    assert "CEO Daily" in registry.product_names


def test_registry_allows_same_name_different_product_type():
    rows = [
        (
            "Fortune 500",
            "Section Front",
            "Daily (web)",
            date(2026, 6, 1),
            "✓",
            "✓",
            "✓",
            "✓",
            "✓",
            "",
            "",
        ),
        (
            "Fortune 500",
            "Fortune List",
            "Annual (launch→EOY)",
            date(2026, 6, 3),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ),
    ]
    reg = InventoryProductRegistry.from_xlsx_bytes(_products_xlsx(rows))
    assert len(reg.rows_for_product("Fortune 500")) == 2
    assert reg.is_inventory_gated("Fortune 500")
    with pytest.raises(ValueError, match="Ambiguous Products row"):
        reg.lookup("Fortune 500")


def test_registry_launch_and_weekdays(registry: InventoryProductRegistry):
    row = registry.lookup("CEO Daily")
    assert row.launch_date == date(2026, 1, 1)
    assert row.runs_on_weekdays == frozenset({"Mon", "Tue", "Wed", "Thu", "Fri"})


def test_availability_blocking_slots(availability: InventoryAvailability):
    blocked = availability.blocking_slots_in_flight(
        "CEO Daily", date(2026, 9, 1), date(2026, 9, 3)
    )
    assert len(blocked) == 1
    assert blocked[0].status == "Sold"
    assert blocked[0].slot_date == date(2026, 9, 2)


def test_not_gated_product_passes(calendar: InventoryCalendar):
    assert calendar.check_inventory_gate(
        "Run of Fortune Display", date(2026, 9, 1), date(2026, 12, 31)
    ) == InventoryGateResult.NOT_GATED
    assert calendar.is_available_for_flight(
        "Branded Content Package", date(2026, 9, 1), date(2026, 12, 31)
    )


def test_gated_blocked_when_sold_in_flight(calendar: InventoryCalendar):
    result = calendar.check_inventory_gate(
        "CEO Daily", date(2026, 9, 1), date(2026, 9, 30)
    )
    assert result == InventoryGateResult.BLOCKED
    assert not calendar.is_available_for_flight(
        "CEO Daily", date(2026, 9, 1), date(2026, 9, 30)
    )


def test_gated_available_when_no_blocking_slots(calendar: InventoryCalendar):
    result = calendar.check_inventory_gate(
        "CEO Daily", date(2026, 9, 1), date(2026, 9, 1)
    )
    assert result == InventoryGateResult.AVAILABLE
    assert calendar.is_available_for_flight(
        "CEO Daily", date(2026, 9, 1), date(2026, 9, 1)
    )


def test_not_launched_before_launch_date(calendar: InventoryCalendar):
    result = calendar.check_inventory_gate(
        "Fortune List", date(2026, 1, 1), date(2026, 3, 31)
    )
    assert result == InventoryGateResult.NOT_LAUNCHED


def test_no_inventory_rows_in_flight(calendar: InventoryCalendar):
    result = calendar.check_inventory_gate(
        "CEO Daily", date(2026, 1, 1), date(2026, 1, 31)
    )
    assert result == InventoryGateResult.NO_INVENTORY_ROWS


def test_holiday_blocks(calendar: InventoryCalendar):
    data = _combined_xlsx(
        _PRODUCT_ROWS,
        [
            (
                date(2026, 11, 1),
                "Sun",
                44,
                "Nov",
                "Newsletter",
                "CEO Daily",
                "Holiday",
                "",
            ),
        ],
    )
    cal = InventoryCalendar.from_xlsx_bytes(data)
    assert cal.check_inventory_gate(
        "CEO Daily", date(2026, 11, 1), date(2026, 11, 1)
    ) == InventoryGateResult.BLOCKED


def test_missing_inventory_sheet_raises():
    wb = Workbook()
    wb.active.title = "Products"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="missing 'Inventory'"):
        InventoryAvailability.from_xlsx_bytes(buf.getvalue())


def test_live_s3_inventory_calendar():
    import os

    import boto3

    bucket = os.environ.get("S3_SNAPSHOT_BUCKET")
    if not bucket:
        pytest.skip("S3_SNAPSHOT_BUCKET not set")
    try:
        cal = load_inventory_calendar_from_s3(boto3.client("s3"), bucket)
    except Exception as exc:
        pytest.skip(f"Inventory calendar not on S3 yet: {exc}")
    assert len(cal.products.product_names) >= 10
    assert cal.products.is_inventory_gated("CEO Daily")
    assert not cal.products.is_inventory_gated("Nonexistent Product XYZ")
