"""Tests for GTM Ideation catalog loaders (Product Category + Product Tags)."""

import io

import pytest
from openpyxl import Workbook

from ingestion.gtm_ideation_catalog import (
    GtmCategoryCatalog,
    GtmIdeationCatalog,
    GtmProductCatalog,
    parse_gtm_tags,
)


def _category_xlsx(rows: list[tuple[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Category"
    ws.append(["Title", "Description"])
    for title, desc in rows:
        ws.append([title, desc])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _product_xlsx(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Tags"
    ws.append(
        ["Product Category", "Product Name", "GTM TAGS", "Deck Path", "Slide #"]
    )
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _combined_xlsx(
    category_rows: list[tuple[str, str]],
    product_rows: list[tuple],
) -> bytes:
    wb = Workbook()
    ws_cat = wb.active
    ws_cat.title = "Product Category"
    ws_cat.append(["Title", "Description"])
    for title, desc in category_rows:
        ws_cat.append([title, desc])
    ws_prod = wb.create_sheet("Product Tags")
    ws_prod.append(
        ["Product Category", "Product Name", "GTM TAGS", "Deck Path", "Slide #"]
    )
    for row in product_rows:
        ws_prod.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_CATEGORY_ROWS = [
    ("Newsletter", "Enter the inbox."),
    ("Print", "Physical print ads."),
]

_PRODUCT_ROWS = [
    (
        "Newsletters",
        "CEO Daily",
        "ceo, c-suite, leadership",
        "Fortune_Newsletters_2026.pptx",
        "3",
    ),
    (
        "Newsletters",
        "Term Sheet",
        "dealmakers, VC, M&A",
        "Fortune_Newsletters_2026.pptx",
        "4",
    ),
    (
        "Vodcasts",
        "Term Sheet",
        "dealmakers, video",
        "Fortune_Premium_Video_2026.pptx",
        "2",
    ),
    (
        "Digital Ads/Programmatic",
        "Crown Unit",
        "high-impact, digital display",
        "Fortune_High_Impact_Media_2026.pptx",
        "5",
    ),
]


@pytest.fixture
def category_catalog() -> GtmCategoryCatalog:
    return GtmCategoryCatalog.from_xlsx_bytes(_category_xlsx(_CATEGORY_ROWS))


@pytest.fixture
def product_catalog() -> GtmProductCatalog:
    return GtmProductCatalog.from_xlsx_bytes(_product_xlsx(_PRODUCT_ROWS))


def test_parse_gtm_tags_splits_and_trims():
    assert parse_gtm_tags("ceo, c-suite, leadership") == (
        "ceo",
        "c-suite",
        "leadership",
    )
    assert parse_gtm_tags("") == ()
    assert parse_gtm_tags("  solo  ") == ("solo",)


def test_category_lookup(category_catalog: GtmCategoryCatalog):
    row = category_catalog.lookup("Newsletter")
    assert row.description == "Enter the inbox."
    assert category_catalog.titles == ["Newsletter", "Print"]


def test_category_missing_raises(category_catalog: GtmCategoryCatalog):
    with pytest.raises(ValueError, match="No Product Category row"):
        category_catalog.lookup("Missing")


def test_product_lookup_by_category(product_catalog: GtmProductCatalog):
    ref = product_catalog.lookup("CEO Daily", "Newsletter")
    assert ref.gtm_tags == ("ceo", "c-suite", "leadership")
    assert ref.slide_number == 3


def test_product_disambiguation(product_catalog: GtmProductCatalog):
    news = product_catalog.lookup("Term Sheet", "Newsletters")
    video = product_catalog.lookup("Term Sheet", "Vodcasts")
    assert news.deck_path.endswith("Newsletters_2026.pptx")
    assert video.deck_path.endswith("Premium_Video_2026.pptx")


def test_products_in_category(product_catalog: GtmProductCatalog):
    rows = product_catalog.products_in_category("Newsletter")
    assert len(rows) == 2
    names = {r.product_name for r in rows}
    assert names == {"CEO Daily", "Term Sheet"}


def test_ambiguous_without_category(product_catalog: GtmProductCatalog):
    with pytest.raises(ValueError, match="Ambiguous Product Tags match"):
        product_catalog.lookup("Term Sheet")


def test_empty_gtm_tags_row_skipped():
    data = _product_xlsx(
        [
            (
                "Newsletters",
                "CEO Daily",
                "ceo",
                "Fortune_Newsletters_2026.pptx",
                "3",
            ),
            ("Newsletters", "Empty Tags", "", "Fortune_Newsletters_2026.pptx", "4"),
        ]
    )
    catalog = GtmProductCatalog.from_xlsx_bytes(data)
    assert catalog.lookup("CEO Daily", "Newsletters").product_name == "CEO Daily"
    with pytest.raises(ValueError, match="No Product Tags row"):
        catalog.lookup("Empty Tags", "Newsletters")


def test_missing_product_category_sheet_raises():
    wb = Workbook()
    wb.active.title = "Product Tags"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="missing 'Product Category'"):
        GtmCategoryCatalog.from_xlsx_bytes(buf.getvalue())


def test_combined_catalog():
    data = _combined_xlsx(_CATEGORY_ROWS, _PRODUCT_ROWS)
    catalog = GtmIdeationCatalog.from_xlsx_bytes(data)
    assert catalog.categories.lookup("Print").title == "Print"
    assert catalog.products.lookup("Crown Unit", "Digital Ads/Programmatic").slide_number == 5


def test_live_s3_workbook_structure():
    """Optional: run against dev bucket when AWS creds + object exist."""
    import os

    import boto3

    bucket = os.environ.get("S3_SNAPSHOT_BUCKET")
    if not bucket:
        pytest.skip("S3_SNAPSHOT_BUCKET not set")
    try:
        catalog = __import__(
            "ingestion.gtm_ideation_catalog", fromlist=["load_gtm_ideation_catalog_from_s3"]
        ).load_gtm_ideation_catalog_from_s3(boto3.client("s3"), bucket)
    except Exception as exc:
        pytest.skip(f"S3 GTM workbook unavailable: {exc}")
    assert len(catalog.categories.titles) >= 5
    assert len(catalog.products.categories) >= 4
    crown = catalog.products.lookup("Crown Unit", "Digital Ads/Programmatic")
    assert "high-impact" in crown.gtm_tags[0] or any(
        "high-impact" in t for t in crown.gtm_tags
    )
