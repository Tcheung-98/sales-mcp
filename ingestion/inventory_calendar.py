"""Inventory calendar loaders for Ideation (PI-2759 Chunk C).

The live workbook splits concerns across two tabs:

* **Products** — master list of inventory-gated placements (name, type, cadence,
  launch date). Products *not* listed here have **no inventory gate** (Logic
  Guide: non-takeover digital, branded content, etc.).
* **Inventory** — dated rows with ``Status`` (Available / Held / Sold / Holiday)
  used to drop candidates when sold or held during the seller flight window.

SOV-specific tabs (Lists Availability, Conference Media Availability) are out
of scope for Chunk C; daily/grid rows on **Inventory** cover most Ideation
candidates.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from enum import Enum

from openpyxl import load_workbook

from ingestion.ideation_data_keys import (
    INVENTORY_COL_DATE,
    INVENTORY_COL_PRODUCT,
    INVENTORY_COL_PRODUCT_TYPE,
    INVENTORY_COL_STATUS,
    INVENTORY_SHEET_INVENTORY,
    INVENTORY_SHEET_PRODUCTS,
    PRODUCTS_COL_CADENCE,
    PRODUCTS_COL_LAUNCH,
    PRODUCTS_COL_PRODUCT,
    PRODUCTS_COL_PRODUCT_TYPE,
    PRODUCTS_DAY_COLUMNS,
    inventory_calendar_s3_key,
)

logger = logging.getLogger(__name__)

_BLOCKING_STATUSES = frozenset({"sold", "held", "holiday"})


class InventoryGateResult(str, Enum):
    """Outcome of an inventory gate check for one product × flight window."""

    NOT_GATED = "not_gated"
    AVAILABLE = "available"
    BLOCKED = "blocked"
    NOT_LAUNCHED = "not_launched"
    NO_INVENTORY_ROWS = "no_inventory_rows"


@dataclass(frozen=True)
class InventoryProductRow:
    """One Products tab row — placements subject to inventory checks."""

    product_name: str
    product_type: str
    cadence: str
    launch_date: date | None
    runs_on_weekdays: frozenset[str]


@dataclass(frozen=True)
class InventorySlot:
    """One Inventory tab row — status for a product on a calendar date."""

    slot_date: date
    product_name: str
    product_type: str
    status: str


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _fold_name(value: str) -> str:
    return value.strip().casefold()


def _parse_excel_date(value: object) -> date | None:
    if value is None or _cell_str(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _normalize_status(raw: object) -> str:
    return _cell_str(raw)


def _is_blocking_status(status: str) -> bool:
    return status.strip().casefold() in _BLOCKING_STATUSES


def _parse_weekday_flags(raw_row: tuple, headers: list[str]) -> frozenset[str]:
    days: set[str] = set()
    for day in PRODUCTS_DAY_COLUMNS:
        if day not in headers:
            continue
        idx = headers.index(day)
        if idx >= len(raw_row):
            continue
        mark = _cell_str(raw_row[idx])
        if mark in {"✓", "x", "X", "Y", "yes", "Yes", "1", "TRUE", "True"}:
            days.add(day)
    return frozenset(days)


def iter_dates_inclusive(start: date, end: date):
    """Yield each calendar date from start through end."""
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


class InventoryProductRegistry:
    """Products tab — which placements have inventory implications."""

    def __init__(self, rows: list[InventoryProductRow]) -> None:
        self._rows = rows
        self._by_name: dict[str, list[InventoryProductRow]] = {}
        for row in rows:
            key = _fold_name(row.product_name)
            bucket = self._by_name.setdefault(key, [])
            for existing in bucket:
                if (
                    existing.product_type == row.product_type
                    and existing.cadence == row.cadence
                ):
                    raise ValueError(
                        f"Duplicate Products row for {row.product_name!r} "
                        f"({row.product_type!r})"
                    )
            bucket.append(row)

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> InventoryProductRegistry:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if INVENTORY_SHEET_PRODUCTS not in wb.sheetnames:
                raise ValueError(
                    f"Inventory calendar missing {INVENTORY_SHEET_PRODUCTS!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[INVENTORY_SHEET_PRODUCTS]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Products sheet is empty") from exc
            headers = [_cell_str(h) for h in header]
            required = {
                PRODUCTS_COL_PRODUCT,
                PRODUCTS_COL_PRODUCT_TYPE,
                PRODUCTS_COL_CADENCE,
                PRODUCTS_COL_LAUNCH,
            }
            missing = required - set(headers)
            if missing:
                raise ValueError(f"Products sheet missing columns: {sorted(missing)}")
            idx = {name: headers.index(name) for name in required}
            parsed: list[InventoryProductRow] = []
            for raw in rows_iter:
                if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
                    continue
                name = _cell_str(raw[idx[PRODUCTS_COL_PRODUCT]])
                if not name:
                    continue
                parsed.append(
                    InventoryProductRow(
                        product_name=name,
                        product_type=_cell_str(raw[idx[PRODUCTS_COL_PRODUCT_TYPE]]),
                        cadence=_cell_str(raw[idx[PRODUCTS_COL_CADENCE]]),
                        launch_date=_parse_excel_date(raw[idx[PRODUCTS_COL_LAUNCH]]),
                        runs_on_weekdays=_parse_weekday_flags(raw, headers),
                    )
                )
        finally:
            wb.close()
        if not parsed:
            raise ValueError("Products sheet has no inventory product rows")
        logger.info("loaded %d Products registry rows", len(parsed))
        return cls(parsed)

    def is_inventory_gated(self, product_name: str) -> bool:
        return bool(self.rows_for_product(product_name))

    def rows_for_product(self, product_name: str) -> list[InventoryProductRow]:
        return list(self._by_name.get(_fold_name(product_name), []))

    def lookup(self, product_name: str) -> InventoryProductRow:
        name = product_name.strip()
        matches = self.rows_for_product(name)
        if not matches:
            raise ValueError(
                f"No Products tab row for placement {name!r} "
                "(exact Product / Placement match required)"
            )
        if len(matches) > 1:
            types = sorted({row.product_type for row in matches})
            raise ValueError(
                f"Ambiguous Products row for {name!r}: "
                f"multiple product types ({types})"
            )
        return matches[0]

    @property
    def product_names(self) -> list[str]:
        return sorted(row.product_name for row in self._rows)


class InventoryAvailability:
    """Inventory tab — dated sold/held/available rows."""

    def __init__(self, slots: list[InventorySlot]) -> None:
        self._slots = list(slots)
        self._by_product: dict[str, list[InventorySlot]] = {}
        for slot in slots:
            self._by_product.setdefault(_fold_name(slot.product_name), []).append(slot)

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> InventoryAvailability:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if INVENTORY_SHEET_INVENTORY not in wb.sheetnames:
                raise ValueError(
                    f"Inventory calendar missing {INVENTORY_SHEET_INVENTORY!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[INVENTORY_SHEET_INVENTORY]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Inventory sheet is empty") from exc
            headers = [_cell_str(h) for h in header]
            required = {
                INVENTORY_COL_DATE,
                INVENTORY_COL_PRODUCT,
                INVENTORY_COL_STATUS,
            }
            missing = required - set(headers)
            if missing:
                raise ValueError(f"Inventory sheet missing columns: {sorted(missing)}")
            idx = {name: headers.index(name) for name in required}
            type_idx = (
                headers.index(INVENTORY_COL_PRODUCT_TYPE)
                if INVENTORY_COL_PRODUCT_TYPE in headers
                else None
            )
            parsed: list[InventorySlot] = []
            for raw in rows_iter:
                if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
                    continue
                slot_date = _parse_excel_date(raw[idx[INVENTORY_COL_DATE]])
                name = _cell_str(raw[idx[INVENTORY_COL_PRODUCT]])
                status = _normalize_status(raw[idx[INVENTORY_COL_STATUS]])
                if slot_date is None or not name or not status:
                    continue
                product_type = (
                    _cell_str(raw[type_idx]) if type_idx is not None else ""
                )
                parsed.append(
                    InventorySlot(
                        slot_date=slot_date,
                        product_name=name,
                        product_type=product_type,
                        status=status,
                    )
                )
        finally:
            wb.close()
        if not parsed:
            raise ValueError("Inventory sheet has no dated rows")
        logger.info("loaded %d Inventory tab rows", len(parsed))
        return cls(parsed)

    def slots_for_product(self, product_name: str) -> list[InventorySlot]:
        return list(self._by_product.get(_fold_name(product_name), []))

    def slots_in_flight(
        self, product_name: str, start: date, end: date
    ) -> list[InventorySlot]:
        if end < start:
            raise ValueError("flight end must be on or after flight start")
        return [
            slot
            for slot in self.slots_for_product(product_name)
            if start <= slot.slot_date <= end
        ]

    def blocking_slots_in_flight(
        self, product_name: str, start: date, end: date
    ) -> list[InventorySlot]:
        return [
            slot
            for slot in self.slots_in_flight(product_name, start, end)
            if _is_blocking_status(slot.status)
        ]


@dataclass(frozen=True)
class InventoryCalendar:
    """Products registry + Inventory grid from one workbook."""

    products: InventoryProductRegistry
    availability: InventoryAvailability

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> InventoryCalendar:
        return cls(
            products=InventoryProductRegistry.from_xlsx_bytes(data),
            availability=InventoryAvailability.from_xlsx_bytes(data),
        )

    def check_inventory_gate(
        self, product_name: str, flight_start: date, flight_end: date
    ) -> InventoryGateResult:
        """Logic Guide availability-first gate for one product × flight window."""
        if flight_end < flight_start:
            raise ValueError("flight_end must be on or after flight_start")

        if not self.products.is_inventory_gated(product_name):
            return InventoryGateResult.NOT_GATED

        launches = [
            row.launch_date
            for row in self.products.rows_for_product(product_name)
            if row.launch_date is not None
        ]
        if launches and flight_end < min(launches):
            return InventoryGateResult.NOT_LAUNCHED

        slots = self.availability.slots_in_flight(
            product_name, flight_start, flight_end
        )
        if not slots:
            return InventoryGateResult.NO_INVENTORY_ROWS

        if self.availability.blocking_slots_in_flight(
            product_name, flight_start, flight_end
        ):
            return InventoryGateResult.BLOCKED

        return InventoryGateResult.AVAILABLE

    def is_available_for_flight(
        self, product_name: str, flight_start: date, flight_end: date
    ) -> bool:
        """True when the product passes the inventory gate (incl. not gated)."""
        result = self.check_inventory_gate(product_name, flight_start, flight_end)
        return result in {
            InventoryGateResult.NOT_GATED,
            InventoryGateResult.AVAILABLE,
        }


def load_inventory_calendar_from_s3(
    s3_client, bucket: str, key: str | None = None
) -> InventoryCalendar:
    """Load inventory calendar from S3 snapshot."""
    s3_key = inventory_calendar_s3_key(key)
    resp = s3_client.get_object(Bucket=bucket, Key=s3_key)
    return InventoryCalendar.from_xlsx_bytes(resp["Body"].read())
