"""Pricing + Benchmarks tab from the inventory calendar (PI-2759 Chunk D).

Rates and commercial terms are **verbatim** from the sheet — never invented.
Logic Guide funding (I2) calls ``lookup`` for a product price and may use
``extract_dollar_amounts`` to pull numeric figures from free-text Pricing cells
(e.g. ``$25,000/day``, ``$60,000``).
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass

from openpyxl import load_workbook

from ingestion.ideation_data_keys import (
    INVENTORY_SHEET_PRICING_BENCHMARKS,
    PRICING_COL_BENCHMARKS,
    PRICING_COL_CONFERENCE_DATE,
    PRICING_COL_EST_IMPS,
    PRICING_COL_LINE_UPDATED,
    PRICING_COL_PRICING,
    PRICING_COL_PRODUCT,
    PRICING_COL_SECTION,
    PRICING_COL_SUBSCRIBERS,
    inventory_calendar_s3_key,
)

logger = logging.getLogger(__name__)

_DOLLAR_PATTERN = re.compile(r"\$[\d,]+(?:\.\d{2})?")


@dataclass(frozen=True)
class PricingRow:
    """One Pricing + Benchmarks row. ``pricing`` is verbatim from the sheet."""

    section: str
    product: str
    pricing: str
    benchmarks: str
    subscribers: str
    est_imps: str
    conference_date: str
    line_last_updated: str


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _fold_name(value: str) -> str:
    return value.strip().casefold()


def extract_dollar_amounts(pricing_text: str) -> tuple[float, ...]:
    """Pull numeric dollar amounts from a Pricing cell (largest first)."""
    amounts: list[float] = []
    for match in _DOLLAR_PATTERN.findall(pricing_text or ""):
        digits = match.lstrip("$").replace(",", "")
        try:
            amounts.append(float(digits))
        except ValueError:
            continue
    return tuple(sorted(set(amounts), reverse=True))


def _header_index(headers: list[str], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


def _is_flat_pricing_header(header: tuple) -> bool:
    headers = [_cell_str(h) for h in header]
    return (
        PRICING_COL_PRODUCT in headers
        and PRICING_COL_PRICING in headers
    )


def _is_formatted_section_header(product_col: str, pricing_col: str) -> bool:
    if not product_col or product_col == PRICING_COL_PRODUCT:
        return False
    if pricing_col and ("$" in pricing_col or pricing_col.startswith("Pricing")):
        return False
    return product_col == product_col.upper() and len(product_col) >= 4


def _parse_flat_pricing_rows(
    rows_iter,
    header: tuple,
) -> list[PricingRow]:
    headers = [_cell_str(h) for h in header]
    required = {PRICING_COL_PRODUCT, PRICING_COL_PRICING}
    missing = required - set(headers)
    if missing:
        raise ValueError(
            f"Pricing + Benchmarks missing columns: {sorted(missing)}"
        )
    idx = {name: headers.index(name) for name in required}
    optional = {
        PRICING_COL_SECTION: "",
        PRICING_COL_BENCHMARKS: "",
        PRICING_COL_SUBSCRIBERS: "",
        PRICING_COL_EST_IMPS: "",
        PRICING_COL_CONFERENCE_DATE: "",
        PRICING_COL_LINE_UPDATED: "",
    }
    opt_idx = {
        name: headers.index(name) if name in headers else None
        for name in optional
    }

    def _opt(raw: tuple, col: str) -> str:
        i = opt_idx[col]
        if i is None or i >= len(raw):
            return optional[col]
        return _cell_str(raw[i])

    parsed: list[PricingRow] = []
    for raw in rows_iter:
        if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
            continue
        product = _cell_str(raw[idx[PRICING_COL_PRODUCT]])
        pricing = _cell_str(raw[idx[PRICING_COL_PRICING]])
        if not product:
            continue
        parsed.append(
            PricingRow(
                section=_opt(raw, PRICING_COL_SECTION),
                product=product,
                pricing=pricing,
                benchmarks=_opt(raw, PRICING_COL_BENCHMARKS),
                subscribers=_opt(raw, PRICING_COL_SUBSCRIBERS),
                est_imps=_opt(raw, PRICING_COL_EST_IMPS),
                conference_date=_opt(raw, PRICING_COL_CONFERENCE_DATE),
                line_last_updated=_opt(raw, PRICING_COL_LINE_UPDATED),
            )
        )
    return parsed


def _parse_formatted_pricing_rows(all_rows: list[tuple]) -> list[PricingRow]:
    """Parse the live SharePoint layout (section blocks in column B)."""
    parsed: list[PricingRow] = []
    current_section = ""
    col_product = 1
    col_pricing = 2
    col_benchmarks: int | None = 3
    col_subscribers: int | None = None
    col_est_imps: int | None = None
    col_line_updated: int | None = None

    for raw in all_rows:
        if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
            continue
        headers = [_cell_str(v) for v in raw]
        product_col = headers[col_product] if len(headers) > col_product else ""
        pricing_col = headers[col_pricing] if len(headers) > col_pricing else ""

        if product_col == PRICING_COL_PRODUCT and pricing_col.startswith("Pricing"):
            col_product = headers.index(PRICING_COL_PRODUCT)
            col_pricing = next(
                i for i, h in enumerate(headers) if h.startswith("Pricing")
            )
            col_benchmarks = _header_index(headers, "Benchmarks", "Benchmarks ")
            col_subscribers = _header_index(headers, "Subscribers")
            col_est_imps = _header_index(headers, "Est. Imps")
            col_line_updated = _header_index(headers, "Last Updated", "Line Last Updated")
            continue

        if _is_formatted_section_header(product_col, pricing_col):
            current_section = product_col
            continue

        if not product_col or product_col == PRICING_COL_PRODUCT:
            continue

        pricing = pricing_col
        if not pricing:
            continue

        def _col(idx: int | None) -> str:
            if idx is None or idx >= len(headers):
                return ""
            return headers[idx]

        parsed.append(
            PricingRow(
                section=current_section,
                product=product_col,
                pricing=pricing,
                benchmarks=_col(col_benchmarks),
                subscribers=_col(col_subscribers),
                est_imps=_col(col_est_imps),
                conference_date="",
                line_last_updated=_col(col_line_updated),
            )
        )
    return parsed


class InventoryPricing:
    """In-memory index of Pricing + Benchmarks rows."""

    def __init__(self, rows: list[PricingRow]) -> None:
        self._rows = list(rows)
        self._by_product: dict[str, list[PricingRow]] = {}
        for row in rows:
            if not row.pricing.strip():
                continue
            self._by_product.setdefault(_fold_name(row.product), []).append(row)

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> InventoryPricing:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if INVENTORY_SHEET_PRICING_BENCHMARKS not in wb.sheetnames:
                raise ValueError(
                    f"Inventory calendar missing "
                    f"{INVENTORY_SHEET_PRICING_BENCHMARKS!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[INVENTORY_SHEET_PRICING_BENCHMARKS]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise ValueError("Pricing + Benchmarks sheet is empty")
            header = all_rows[0]
            if _is_flat_pricing_header(header):
                parsed = _parse_flat_pricing_rows(iter(all_rows[1:]), header)
            else:
                parsed = _parse_formatted_pricing_rows(all_rows)
        finally:
            wb.close()
        catalog = cls(parsed)
        if not catalog._rows:
            raise ValueError("Pricing + Benchmarks sheet has no product rows")
        if not catalog._by_product:
            raise ValueError(
                "Pricing + Benchmarks sheet has no rows with non-empty Pricing"
            )
        logger.info(
            "loaded %d Pricing rows (%d priced products)",
            len(parsed),
            len(catalog._by_product),
        )
        return catalog

    @property
    def product_names(self) -> list[str]:
        return sorted({row.product for row in self._rows if row.pricing.strip()})

    def lookup(self, product_name: str) -> PricingRow:
        """Exact Product match (case-insensitive). Fail loud if missing or unpriced."""
        name = product_name.strip()
        candidates = self._by_product.get(_fold_name(name), [])
        if not candidates:
            raise ValueError(
                f"No Pricing + Benchmarks row for product {name!r} "
                "(exact Product match required)"
            )
        if len(candidates) > 1:
            sections = sorted({c.section for c in candidates})
            raise ValueError(
                f"Ambiguous Pricing + Benchmarks match for product {name!r}: "
                f"{len(candidates)} rows (sections: {sections})"
            )
        return candidates[0]

    def primary_amount(self, product_name: str) -> float:
        """Largest dollar figure parsed from the product's Pricing cell."""
        row = self.lookup(product_name)
        amounts = extract_dollar_amounts(row.pricing)
        if not amounts:
            raise ValueError(
                f"Pricing row for {product_name!r} has no parseable dollar amount: "
                f"{row.pricing!r}"
            )
        return amounts[0]


def load_inventory_pricing_from_s3(
    s3_client, bucket: str, key: str | None = None
) -> InventoryPricing:
    """Load pricing from the inventory calendar S3 snapshot."""
    s3_key = inventory_calendar_s3_key(key)
    resp = s3_client.get_object(Bucket=bucket, Key=s3_key)
    return InventoryPricing.from_xlsx_bytes(resp["Body"].read())
