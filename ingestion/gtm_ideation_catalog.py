"""Ideation catalog from Fortune_AITool_GTM_Database (PI-2759 Chunk B).

Loads Product Category (display names + descriptions) and Product Tags rows
for Logic Guide candidate pools. GTM tag strings live on the Product Tags
sheet in column ``GTM TAGS`` — there is no separate ``GTM Tags`` tab in the
live workbook.

Deck Path / Slide # on Product Tags are loaded for cross-reference but exact
clone lookup remains ``gtm_product_map.GtmProductMap`` (A5).
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass

from openpyxl import load_workbook

from ingestion.gtm_product_map import normalize_category
from ingestion.ideation_data_keys import (
    GTM_COL_CATEGORY_DESCRIPTION,
    GTM_COL_CATEGORY_TITLE,
    GTM_COL_GTM_TAGS,
    GTM_COL_PRODUCT_CATEGORY,
    GTM_COL_PRODUCT_NAME,
    GTM_SHEET_PRODUCT_CATEGORY,
    GTM_SHEET_PRODUCT_TAGS,
    gtm_database_s3_key,
)

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class GtmCategoryRow:
    """One Product Category sheet row (marketing category title + blurb)."""

    title: str
    description: str


@dataclass(frozen=True)
class GtmProductCandidate:
    """One Product Tags row for Ideation — name, category, parsed GTM tags."""

    product_name: str
    category: str
    gtm_tags: tuple[str, ...]
    deck_path: str
    slide_number: int | None


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_gtm_tags(raw: object) -> tuple[str, ...]:
    """Split comma-separated GTM TAGS cell into normalized tag strings."""
    text = _cell_str(raw)
    if not text:
        return ()
    parts = [part.strip() for part in text.split(",")]
    return tuple(part for part in parts if part)


def _parse_optional_slide(raw: object) -> int | None:
    text = _cell_str(raw)
    if not text:
        return None
    if text.isdigit():
        n = int(text)
        return n if n >= 1 else None
    return None


class GtmCategoryCatalog:
    """In-memory index of Product Category sheet rows."""

    def __init__(self, rows: list[GtmCategoryRow]) -> None:
        self._rows = list(rows)
        self._by_title: dict[str, GtmCategoryRow] = {}
        for row in rows:
            key = row.title.casefold()
            existing = self._by_title.get(key)
            if existing is not None and existing.description != row.description:
                raise ValueError(
                    f"Duplicate Product Category title {row.title!r} with "
                    f"conflicting descriptions"
                )
            self._by_title[key] = row

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> GtmCategoryCatalog:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if GTM_SHEET_PRODUCT_CATEGORY not in wb.sheetnames:
                raise ValueError(
                    f"GTM database missing {GTM_SHEET_PRODUCT_CATEGORY!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[GTM_SHEET_PRODUCT_CATEGORY]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Product Category sheet is empty") from exc
            headers = [_cell_str(h) for h in header]
            required = {GTM_COL_CATEGORY_TITLE, GTM_COL_CATEGORY_DESCRIPTION}
            missing = required - set(headers)
            if missing:
                raise ValueError(
                    f"Product Category missing columns: {sorted(missing)}"
                )
            idx = {name: headers.index(name) for name in required}
            parsed: list[GtmCategoryRow] = []
            for raw in rows_iter:
                if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
                    continue
                title = _cell_str(raw[idx[GTM_COL_CATEGORY_TITLE]])
                if not title:
                    continue
                description = _cell_str(raw[idx[GTM_COL_CATEGORY_DESCRIPTION]])
                parsed.append(GtmCategoryRow(title=title, description=description))
        finally:
            wb.close()
        if not parsed:
            raise ValueError("Product Category sheet has no category rows")
        logger.info("loaded %d Product Category rows", len(parsed))
        return cls(parsed)

    @property
    def titles(self) -> list[str]:
        return [row.title for row in self._rows]

    def lookup(self, title: str) -> GtmCategoryRow:
        name = title.strip()
        row = self._by_title.get(name.casefold())
        if row is None:
            raise ValueError(
                f"No Product Category row for title {name!r} "
                "(exact Title match required)"
            )
        return row


class GtmProductCatalog:
    """In-memory index of Product Tags rows for Ideation (GTM TAGS candidate pool)."""

    def __init__(self, rows: list[GtmProductCandidate]) -> None:
        self._rows = list(rows)
        self._by_name: dict[str, list[GtmProductCandidate]] = {}
        self._by_category: dict[str, list[GtmProductCandidate]] = {}
        for row in rows:
            self._by_name.setdefault(row.product_name, []).append(row)
            self._by_category.setdefault(row.category, []).append(row)

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> GtmProductCatalog:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if GTM_SHEET_PRODUCT_TAGS not in wb.sheetnames:
                raise ValueError(
                    f"GTM database missing {GTM_SHEET_PRODUCT_TAGS!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[GTM_SHEET_PRODUCT_TAGS]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Product Tags sheet is empty") from exc
            headers = [_cell_str(h) for h in header]
            required = {
                GTM_COL_PRODUCT_CATEGORY,
                GTM_COL_PRODUCT_NAME,
                GTM_COL_GTM_TAGS,
            }
            missing = required - set(headers)
            if missing:
                raise ValueError(
                    f"Product Tags missing columns for Ideation catalog: "
                    f"{sorted(missing)}"
                )
            idx = {name: headers.index(name) for name in required}
            deck_idx = (
                headers.index("Deck Path") if "Deck Path" in headers else None
            )
            slide_idx = headers.index("Slide #") if "Slide #" in headers else None
            parsed: list[GtmProductCandidate] = []
            for raw in rows_iter:
                if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
                    continue
                name = _cell_str(raw[idx[GTM_COL_PRODUCT_NAME]])
                if not name:
                    continue
                category = _cell_str(raw[idx[GTM_COL_PRODUCT_CATEGORY]])
                tags = parse_gtm_tags(raw[idx[GTM_COL_GTM_TAGS]])
                if not tags:
                    logger.warning(
                        "skipping Product Tags row %r: empty GTM TAGS", name
                    )
                    continue
                deck_path = (
                    _cell_str(raw[deck_idx]) if deck_idx is not None else ""
                )
                slide_number = (
                    _parse_optional_slide(raw[slide_idx])
                    if slide_idx is not None
                    else None
                )
                parsed.append(
                    GtmProductCandidate(
                        product_name=name,
                        category=category,
                        gtm_tags=tags,
                        deck_path=deck_path,
                        slide_number=slide_number,
                    )
                )
        finally:
            wb.close()
        if not parsed:
            raise ValueError("Product Tags sheet has no Ideation product rows")
        logger.info("loaded %d Product Tags Ideation rows", len(parsed))
        return cls(parsed)

    @property
    def product_names(self) -> list[str]:
        return sorted(self._by_name)

    @property
    def categories(self) -> list[str]:
        return sorted(self._by_category)

    def products_in_category(self, category: str) -> list[GtmProductCandidate]:
        wanted = normalize_category(category)
        if wanted in self._by_category:
            return list(self._by_category[wanted])
        exact = self._by_category.get(category.strip())
        return list(exact) if exact else []

    def lookup(self, product_name: str, category: str | None = None) -> GtmProductCandidate:
        name = product_name.strip()
        candidates = self._by_name.get(name, [])
        if not candidates:
            raise ValueError(
                f"No Product Tags row for product {name!r} "
                "(exact Product Name match required)"
            )
        if category:
            wanted = normalize_category(category)
            matched = [c for c in candidates if c.category == wanted]
            if len(matched) == 1:
                return matched[0]
            if not matched:
                available = sorted({c.category for c in candidates})
                raise ValueError(
                    f"No Product Tags row for product {name!r} "
                    f"in category {wanted!r} (available: {available})"
                )
            raise ValueError(
                f"Ambiguous Product Tags match for product {name!r} "
                f"in category {wanted!r}: {len(matched)} rows"
            )
        if len(candidates) == 1:
            return candidates[0]
        cats = sorted({c.category for c in candidates})
        raise ValueError(
            f"Ambiguous Product Tags match for product {name!r}: "
            f"multiple categories {cats}; pass category to disambiguate"
        )


@dataclass(frozen=True)
class GtmIdeationCatalog:
    """Product Category + Product Tags loaded from one GTM workbook."""

    categories: GtmCategoryCatalog
    products: GtmProductCatalog

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> GtmIdeationCatalog:
        return cls(
            categories=GtmCategoryCatalog.from_xlsx_bytes(data),
            products=GtmProductCatalog.from_xlsx_bytes(data),
        )


def load_gtm_ideation_catalog_from_s3(
    s3_client, bucket: str, key: str | None = None
) -> GtmIdeationCatalog:
    """Load Ideation GTM catalog from the synced GTM workbook on S3."""
    s3_key = gtm_database_s3_key(key)
    resp = s3_client.get_object(Bucket=bucket, Key=s3_key)
    return GtmIdeationCatalog.from_xlsx_bytes(resp["Body"].read())
