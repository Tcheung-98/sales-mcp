"""Audience Data tab from Fortune_AITool_GTM_Database (same xlsx as A5).

Reach and Index are Data Pulls — verbatim from the sheet, never invented.
Matching targeting_details to known segment names is deterministic string
match, not Claude.
"""

from __future__ import annotations

import io
import logging
import os
from dataclasses import dataclass

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_AUDIENCE_SHEET = "Audience Data"
_COL_SEGMENT = "Audience Segment"
_COL_REACH = "Reach"
_COL_INDEX = "Index"
_DEFAULT_GTM_DATABASE_KEY = "templates/Fortune_AITool_GTM_Database.xlsx"


@dataclass(frozen=True)
class AudienceRow:
    """One Audience Data row. reach/index are display strings from the sheet."""

    segment: str
    reach: str
    index: str


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _index_sort_key(row: AudienceRow) -> int:
    try:
        return int(row.index)
    except ValueError:
        return 0


def match_segment_names(targeting_details: str, known_names: list[str]) -> list[str]:
    """Return known Audience Data names found in targeting text.

    Case-insensitive substring match. Unrecognized prose is ignored. If both a
    short name and a longer name that contains it match, keep only the longer
    name. Order is first appearance in ``targeting_details``.
    """
    text = targeting_details or ""
    text_l = text.lower()
    hits: list[str] = []
    for name in known_names:
        if name.lower() in text_l:
            hits.append(name)
    filtered = [
        name
        for name in hits
        if not any(
            name != other and name.lower() in other.lower() for other in hits
        )
    ]
    filtered.sort(key=lambda n: text_l.find(n.lower()))
    return filtered


class AudienceData:
    """In-memory index of Audience Data rows."""

    def __init__(self, rows: list[AudienceRow]) -> None:
        unique: list[AudienceRow] = []
        self._by_lower: dict[str, AudienceRow] = {}
        for row in rows:
            key = row.segment.lower()
            existing = self._by_lower.get(key)
            if existing is None:
                self._by_lower[key] = row
                unique.append(row)
                continue
            if existing.reach != row.reach or existing.index != row.index:
                raise ValueError(
                    f"Duplicate Audience Data segment {row.segment!r} with "
                    f"conflicting Reach/Index"
                )
        self._rows = unique

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> AudienceData:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if _AUDIENCE_SHEET not in wb.sheetnames:
                raise ValueError(
                    f"GTM database missing {_AUDIENCE_SHEET!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[_AUDIENCE_SHEET]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Audience Data sheet is empty") from exc
            headers = [_cell_str(h) for h in header]
            required = {_COL_SEGMENT, _COL_REACH, _COL_INDEX}
            missing = required - set(headers)
            if missing:
                raise ValueError(
                    f"Audience Data missing columns: {sorted(missing)}"
                )
            idx = {name: headers.index(name) for name in required}
            parsed: list[AudienceRow] = []
            for raw in rows_iter:
                if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
                    continue
                segment = _cell_str(raw[idx[_COL_SEGMENT]])
                if not segment:
                    continue
                reach = _cell_str(raw[idx[_COL_REACH]])
                index = _cell_str(raw[idx[_COL_INDEX]])
                if not reach or not index:
                    logger.warning(
                        "skipping Audience Data row %r: empty Reach or Index",
                        segment,
                    )
                    continue
                parsed.append(AudienceRow(segment=segment, reach=reach, index=index))
        finally:
            wb.close()
        if not parsed:
            raise ValueError("Audience Data sheet has no audience rows")
        logger.info("loaded %d Audience Data rows", len(parsed))
        return cls(parsed)

    @property
    def segment_names(self) -> list[str]:
        return [row.segment for row in self._rows]

    def lookup(self, segment: str) -> AudienceRow:
        """Exact segment match (case-insensitive). Fail loud if missing."""
        name = segment.strip()
        row = self._by_lower.get(name.lower())
        if row is None:
            raise ValueError(
                f"No Audience Data row for segment {name!r} "
                "(exact Audience Segment match required)"
            )
        return row

    def match_targeting(self, targeting_details: str) -> list[AudienceRow]:
        """Map targeting_details to Audience Data rows (no invented metrics)."""
        names = match_segment_names(targeting_details, self.segment_names)
        return [self.lookup(name) for name in names]


def rank_segments_by_index(rows: list[AudienceRow]) -> list[AudienceRow]:
    """Highest Index first (for >6 truncate). Does not invent Index values."""
    return sorted(rows, key=_index_sort_key, reverse=True)


def load_audience_data_from_s3(
    s3_client, bucket: str, key: str | None = None
) -> AudienceData:
    """Load Audience Data from the same GTM xlsx A5 uses."""
    s3_key = key or os.environ.get("GTM_DATABASE_KEY", _DEFAULT_GTM_DATABASE_KEY)
    resp = s3_client.get_object(Bucket=bucket, Key=s3_key)
    return AudienceData.from_xlsx_bytes(resp["Body"].read())
