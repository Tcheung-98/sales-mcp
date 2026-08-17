"""Exact product → Deck Path / Slide # map from Fortune_AITool_GTM_Database.

Source of truth: Product Tags sheet (not Titan similarity). Lookup is exact on
Product Name, disambiguated with Product Category when names collide.
"""

from __future__ import annotations

import io
import logging
import os
import re
from dataclasses import dataclass

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_PRODUCT_TAGS_SHEET = "Product Tags"
_COL_CATEGORY = "Product Category"
_COL_NAME = "Product Name"
_COL_DECK_PATH = "Deck Path"
_COL_SLIDE = "Slide #"

# Schema / preferred-platform strings → GTM Product Tags "Product Category"
_CATEGORY_ALIASES: dict[str, str] = {
    "Newsletter": "Newsletters",
    "Newsletters": "Newsletters",
    "Digital Media": "Digital Ads/Programmatic",
    "Digital Ads/Programmatic": "Digital Ads/Programmatic",
    "Branded Content": "Branded Content",
    "Vodcasts": "Vodcasts",
    "Print": "Print",
    "Events": "Events",
    "Conference Sponsorship/Media": "Events",
    "Lists & Rankings Sponsorship": "Lists & Rankings Sponsorship",
}

_DEFAULT_GTM_DATABASE_KEY = "templates/Fortune_AITool_GTM_Database.xlsx"
_DEFAULT_PRODUCT_DECKS_PREFIX = "product-decks/"


@dataclass(frozen=True)
class ProductSlideRef:
    """Exact slide coordinates for one funded product."""

    product_name: str
    category: str
    deck_path: str
    slide_number: int


def normalize_category(category: str) -> str:
    """Map schema / platform category strings onto GTM Product Tags values."""
    key = category.strip()
    return _CATEGORY_ALIASES.get(key, key)


def normalize_deck_filename(deck_path: str) -> str:
    """Bare Hunter filename → S3 object basename (ensure .pptx)."""
    name = deck_path.strip()
    if not name:
        raise ValueError("Deck Path is empty")
    if not name.lower().endswith(".pptx"):
        name = f"{name}.pptx"
    return name


def product_deck_s3_key(
    deck_path: str, prefix: str | None = None
) -> str:
    """S3 key for a Fortune Hunter product deck cached under product-decks/."""
    base = prefix if prefix is not None else os.environ.get(
        "PRODUCT_DECKS_PREFIX", _DEFAULT_PRODUCT_DECKS_PREFIX
    )
    if not base.endswith("/"):
        base = f"{base}/"
    return f"{base}{normalize_deck_filename(deck_path)}"


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_slide_number(raw: object, *, product_name: str) -> int:
    text = _cell_str(raw)
    if not text:
        raise ValueError(f"Product Tags row for {product_name!r} has empty Slide #")
    match = re.fullmatch(r"(\d+)", text)
    if not match:
        raise ValueError(
            f"Product Tags row for {product_name!r} has invalid Slide #: {raw!r}"
        )
    n = int(match.group(1))
    if n < 1:
        raise ValueError(
            f"Product Tags row for {product_name!r} has non-positive Slide #: {n}"
        )
    return n


class GtmProductMap:
    """In-memory index of Product Tags rows."""

    def __init__(self, rows: list[ProductSlideRef]) -> None:
        self._rows = list(rows)
        self._by_name: dict[str, list[ProductSlideRef]] = {}
        for row in self._rows:
            self._by_name.setdefault(row.product_name, []).append(row)

    @classmethod
    def from_xlsx_bytes(cls, data: bytes) -> GtmProductMap:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            if _PRODUCT_TAGS_SHEET not in wb.sheetnames:
                raise ValueError(
                    f"GTM database missing {_PRODUCT_TAGS_SHEET!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[_PRODUCT_TAGS_SHEET]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Product Tags sheet is empty") from exc
            headers = [_cell_str(h) for h in header]
            required = {_COL_CATEGORY, _COL_NAME, _COL_DECK_PATH, _COL_SLIDE}
            missing = required - set(headers)
            if missing:
                raise ValueError(
                    f"Product Tags missing columns: {sorted(missing)}"
                )
            idx = {name: headers.index(name) for name in required}
            parsed: list[ProductSlideRef] = []
            for raw in rows_iter:
                if raw is None or all(v is None or _cell_str(v) == "" for v in raw):
                    continue
                name = _cell_str(raw[idx[_COL_NAME]])
                if not name:
                    continue
                category = _cell_str(raw[idx[_COL_CATEGORY]])
                deck_path = _cell_str(raw[idx[_COL_DECK_PATH]])
                if not deck_path:
                    raise ValueError(
                        f"Product Tags row for {name!r} has empty Deck Path"
                    )
                slide_number = _parse_slide_number(
                    raw[idx[_COL_SLIDE]], product_name=name
                )
                parsed.append(
                    ProductSlideRef(
                        product_name=name,
                        category=category,
                        deck_path=deck_path,
                        slide_number=slide_number,
                    )
                )
        finally:
            wb.close()
        if not parsed:
            raise ValueError("Product Tags sheet has no product rows")
        logger.info("loaded %d Product Tags rows", len(parsed))
        return cls(parsed)

    def lookup(self, product_name: str, category: str | None = None) -> ProductSlideRef:
        """Resolve exact Product Name (+ category when the name is ambiguous)."""
        name = product_name.strip()
        candidates = self._by_name.get(name, [])
        if not candidates:
            raise ValueError(
                f"No GTM Product Tags row for product {name!r} "
                "(exact Product Name match required)"
            )

        # Deduplicate identical Deck Path + Slide # rows (duplicate tag rows).
        unique: dict[tuple[str, int, str], ProductSlideRef] = {}
        for row in candidates:
            key = (row.deck_path, row.slide_number, row.category)
            unique[key] = row
        candidates = list(unique.values())

        if category:
            wanted = normalize_category(category)
            matched = [c for c in candidates if c.category == wanted]
            if len(matched) == 1:
                return matched[0]
            if not matched:
                available = sorted({c.category for c in candidates})
                raise ValueError(
                    f"No GTM Product Tags row for product {name!r} "
                    f"in category {wanted!r} (available: {available})"
                )
            coords = sorted({(m.deck_path, m.slide_number) for m in matched})
            raise ValueError(
                f"Ambiguous GTM Product Tags match for product {name!r} "
                f"in category {wanted!r}: multiple Deck Path/Slide # {coords}"
            )

        if len(candidates) == 1:
            return candidates[0]

        cats = sorted({c.category for c in candidates})
        raise ValueError(
            f"Ambiguous GTM Product Tags match for product {name!r}: "
            f"multiple categories {cats}; pass category to disambiguate"
        )


def load_gtm_product_map_from_s3(s3_client, bucket: str, key: str | None = None) -> GtmProductMap:
    """Load the GTM xlsx from S3 (synced from Fortune Hunter)."""
    s3_key = key or os.environ.get("GTM_DATABASE_KEY", _DEFAULT_GTM_DATABASE_KEY)
    resp = s3_client.get_object(Bucket=bucket, Key=s3_key)
    return GtmProductMap.from_xlsx_bytes(resp["Body"].read())
